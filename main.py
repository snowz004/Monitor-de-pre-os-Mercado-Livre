"""
Monitor de preço Mercado Livre: baixa a página, extrai o preço, registra no Excel
e envia e-mail quando o valor fica abaixo do limite configurado no .env.
"""

from __future__ import annotations

import json
import os
import re
import smtplib
from datetime import datetime
from email.message import EmailMessage
from typing import Optional
from urllib.parse import urlparse, urlunparse

import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

# ---------------------------------------------------------------------------
# Bloco: URL fixa do produto (normalizada sem fragmentos de tracking).
# O fragmento (#...) não é enviado ao servidor; mantemos a base limpa para o GET.
# ---------------------------------------------------------------------------
PRODUCT_URL = (
    "https://www.mercadolivre.com.br/motinha-eletrica-scooter-bike-500w-encosto-carona-gts/up/MLBU3777079997"
)


def _normalize_url(url: str) -> str:
    """Remove fragmento (#...) da URL para requisição HTTP."""
    parsed = urlparse(url.strip())
    return urlunparse(parsed._replace(fragment=""))


# ---------------------------------------------------------------------------
# Bloco: carrega variáveis de ambiente do arquivo .env (python-dotenv).
# SMTP_* e credenciais ficam fora do código por segurança.
# ---------------------------------------------------------------------------
load_dotenv()

EMAIL_TO = os.getenv("EMAIL_TO", "").strip()
PRICE_THRESHOLD_STR = os.getenv("PRICE_THRESHOLD", "").strip()
SMTP_HOST = os.getenv("SMTP_HOST", "").strip()
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "").strip()
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "").strip()
EMAIL_FROM = os.getenv("EMAIL_FROM", SMTP_USER).strip()
EXCEL_PATH = os.getenv("EXCEL_PATH", "consultas_preco.xlsx").strip()


def _parse_threshold() -> Optional[float]:
    """Converte PRICE_THRESHOLD do .env para float (vírgula ou ponto)."""
    if not PRICE_THRESHOLD_STR:
        return None
    normalized = PRICE_THRESHOLD_STR.replace(",", ".")
    return float(normalized)


# ---------------------------------------------------------------------------
# Bloco: download da página HTML com User-Agent de navegador.
# Alguns sites bloqueiam clientes sem cabeçalho realista.
# ---------------------------------------------------------------------------
def fetch_html(url: str) -> str:
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
    }
    response = requests.get(url, headers=headers, timeout=30)
    response.raise_for_status()
    return response.text


# ---------------------------------------------------------------------------
# Bloco: extração do preço com BeautifulSoup — várias estratégias usuais no ML.
# Ordem: meta itemprop/og → JSON-LD → spans andes-money-amount → regex em scripts.
# ---------------------------------------------------------------------------
def parse_price_brl(html: str) -> float:
    soup = BeautifulSoup(html, "html.parser")

    # Meta tags (SEO / schema)
    for selector in [
        ('meta', {'itemprop': 'price'}),
        ('meta', {'property': 'product:price:amount'}),
    ]:
        tag = soup.find(*selector)
        if tag and tag.get("content"):
            return float(str(tag["content"]).replace(",", "."))

    # JSON-LD (muitas páginas de produto embutem oferta aqui)
    for script in soup.find_all("script", type="application/ld+json"):
        raw = script.string or script.get_text() or ""
        raw = raw.strip()
        if not raw:
            continue
        try:
            data = json.loads(raw)
        except json.JSONDecodeError:
            continue
        candidates = data if isinstance(data, list) else [data]
        for item in candidates:
            if not isinstance(item, dict):
                continue
            offers = item.get("offers")
            if isinstance(offers, dict) and "price" in offers:
                return float(str(offers["price"]).replace(",", "."))
            if isinstance(offers, list) and offers and isinstance(offers[0], dict):
                p = offers[0].get("price")
                if p is not None:
                    return float(str(p).replace(",", "."))

    # Componente visual andes-money-amount (parte inteira + centavos)
    fraction = soup.select_one(".andes-money-amount__fraction")
    if fraction:
        cents_el = soup.select_one(".andes-money-amount__cents")
        whole = re.sub(r"\D", "", fraction.get_text())
        cents = re.sub(r"\D", "", cents_el.get_text()) if cents_el else "00"
        if whole:
            return float(f"{whole}.{cents.zfill(2)[:2]}")

    # Fallback: primeiro número com formato de preço BRL no HTML
    text = soup.get_text(" ", strip=True)
    match = re.search(
        r"R\$\s*([\d]{1,3}(?:\.[\d]{3})*,\d{2}|[\d]+,\d{2})",
        text,
        re.IGNORECASE,
    )
    if match:
        num = match.group(1).replace(".", "").replace(",", ".")
        return float(num)

    raise ValueError(
        "Não foi possível encontrar o preço na página. "
        "O layout do Mercado Livre pode ter mudado ou a página exige JavaScript."
    )


# ---------------------------------------------------------------------------
# Bloco: persistência em Excel com openpyxl — uma linha por consulta.
# Se o arquivo não existir, cria planilha com cabeçalho.
# ---------------------------------------------------------------------------
def append_excel_row(path: str, consulted_at: datetime, price: float, source_url: str) -> None:
    if os.path.isfile(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Consultas"
        ws.append(["Data/Hora da consulta", "Preço (BRL)", "URL"])

    ws.append(
        [
            consulted_at.strftime("%Y-%m-%d %H:%M:%S"),
            price,
            source_url,
        ]
    )
    wb.save(path)


# ---------------------------------------------------------------------------
# Bloco: envio de e-mail via SMTP (TLS na porta 587 é o caso mais comum).
# Disparado só quando price < PRICE_THRESHOLD e as variáveis SMTP estão ok.
# ---------------------------------------------------------------------------
def send_price_alert(
    to_addr: str,
    price: float,
    threshold: float,
    product_url: str,
) -> None:
    if not all([SMTP_HOST, SMTP_USER, SMTP_PASSWORD, EMAIL_FROM, to_addr]):
        raise RuntimeError(
            "Configure SMTP_HOST, SMTP_USER, SMTP_PASSWORD, EMAIL_FROM e EMAIL_TO no .env "
            "para enviar o alerta."
        )

    msg = EmailMessage()
    msg["Subject"] = f"Alerta: preço R$ {price:.2f} abaixo do limite R$ {threshold:.2f}"
    msg["From"] = EMAIL_FROM
    msg["To"] = to_addr
    msg.set_content(
        f"O preço atual do produto é R$ {price:.2f}.\n"
        f"Seu limite configurado é R$ {threshold:.2f}.\n\n"
        f"Link: {product_url}\n"
    )

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
        smtp.starttls()
        smtp.login(SMTP_USER, SMTP_PASSWORD)
        smtp.send_message(msg)


# ---------------------------------------------------------------------------
# Bloco: fluxo principal — busca, parse, grava Excel, avalia limite e opcionalmente e-mail.
# ---------------------------------------------------------------------------
def main() -> None:
    url = _normalize_url(PRODUCT_URL)
    consulted_at = datetime.now()

    html = fetch_html(url)
    price = parse_price_brl(html)

    append_excel_row(EXCEL_PATH, consulted_at, price, url)

    threshold = _parse_threshold()
    if threshold is not None and price < threshold:
        if not EMAIL_TO:
            raise RuntimeError(
                "PRICE_THRESHOLD definido mas EMAIL_TO está vazio. "
                "Defina o destinatário no .env para receber o alerta."
            )
        send_price_alert(EMAIL_TO, price, threshold, url)
        print(
            f"Preço R$ {price:.2f} abaixo do limite R$ {threshold:.2f}. "
            f"E-mail enviado para {EMAIL_TO}."
        )
    else:
        if threshold is None:
            print(f"Preço atual: R$ {price:.2f}. Limite não configurado — sem e-mail.")
        else:
            print(
                f"Preço atual: R$ {price:.2f} (limite R$ {threshold:.2f}). "
                "Sem alerta."
            )
    print(f"Registro salvo em: {os.path.abspath(EXCEL_PATH)}")


if __name__ == "__main__":
    main()
